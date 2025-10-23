#!/usr/bin/env python3
"""
Скрипт для автоматического воссоздания Excel файла "Копия 123 Миша.xlsx"
на основе данных из Django базы данных
"""

import os
import sys
import django
from datetime import datetime

# Настройка Django
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'wb_analytics_project.settings')
django.setup()

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from analytics.models import Nomenclature, FinancialReport, SalesAnalysis, StockBalance, SummaryData

def create_excel_structure():
    """Создает базовую структуру Excel файла"""
    print("📊 Создание базовой структуры Excel файла...")
    
    # Создаем новую книгу
    wb = openpyxl.Workbook()
    
    # Удаляем стандартный лист
    wb.remove(wb.active)
    
    # Создаем все листы в правильном порядке
    sheet_names = [
        'Номенклатуры (вставить)',
        'Финотчеты (вставить)', 
        'Остатки (вставить)',
        'Анализ продаж',
        'Сводный',
        'План по выкупам',
        'Лист10',
        'Лист11'
    ]
    
    for sheet_name in sheet_names:
        wb.create_sheet(sheet_name)
        print(f"  ✅ Создан лист: {sheet_name}")
    
    return wb

def setup_nomenclature_sheet(wb):
    """Настраивает лист 'Номенклатуры (вставить)'"""
    print("📋 Настройка листа 'Номенклатуры (вставить)'...")
    
    ws = wb['Номенклатуры (вставить)']
    
    # Заголовки столбцов
    headers = [
        'Бренд', 'Предмет', 'Код размера (chrt_id)', 'Артикул продавца',
        'Артикул WB', 'Размер', 'Баркод', 'Комплектация', 'Состав',
        'Себестоимость', 'B025425BE13000C53FA9'
    ]
    
    # Записываем заголовки
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Настройка ширины столбцов
    column_widths = [15, 20, 20, 15, 15, 10, 15, 12, 30, 12, 20]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # Загружаем данные из базы
    nomenclatures = Nomenclature.objects.all()
    for idx, nom in enumerate(nomenclatures, 2):
        ws.cell(row=idx, column=1, value=nom.brand)
        ws.cell(row=idx, column=2, value=nom.subject)
        ws.cell(row=idx, column=3, value=nom.size_code)
        ws.cell(row=idx, column=4, value=nom.supplier_article)
        ws.cell(row=idx, column=5, value=nom.wb_article)
        ws.cell(row=idx, column=6, value=nom.size)
        ws.cell(row=idx, column=7, value=nom.barcode)
        ws.cell(row=idx, column=8, value=nom.equipment)
        ws.cell(row=idx, column=9, value=nom.composition)
        ws.cell(row=idx, column=10, value=nom.cost_price)
        ws.cell(row=idx, column=11, value=None)  # B025425BE13000C53FA9
    
    print(f"  ✅ Загружено {len(nomenclatures)} записей номенклатур")

def setup_financial_reports_sheet(wb):
    """Настраивает лист 'Финотчеты (вставить)'"""
    print("💰 Настройка листа 'Финотчеты (вставить)'...")
    
    ws = wb['Финотчеты (вставить)']
    
    # Заголовки столбцов (71 столбец)
    headers = [
        '№', 'Номер поставки', 'Предмет', 'Код номенклатуры', 'Бренд',
        'Артикул поставщика', 'Название', 'Размер', 'Баркод', 'Тип документа',
        'Обоснование для оплаты', 'Дата заказа покупателем', 'Дата продажи',
        'Кол-во', 'Цена розничная', 'Вайлдберриз реализовал Товар (Пр)',
        'Согласованный продуктовый дисконт, %', 'Промокод %',
        'Итоговая согласованная скидка', 'Цена розничная с учетом согласованной скидки',
        'Размер снижения кВВ из-за рейтинга, %', 'Размер снижения кВВ из-за акции, %',
        'Скидка постоянного Покупателя (СПП)', 'Размер кВВ, %',
        'Размер кВВ без НДС, % Базовый', 'Итоговый кВВ без НДС, %',
        'Вознаграждение с продаж до вычета услуг поверенного, без НДС',
        'Возмещение за выдачу и возврат товаров на ПВЗ',
        'Эквайринг/Комиссии за организацию платежей',
        'Размер комиссии за эквайринг/Комиссии за организацию платежей, %',
        'Тип платежа за Эквайринг/Комиссии за организацию платежей',
        'Вознаграждение Вайлдберриз (ВВ), без НДС',
        'НДС с Вознаграждения Вайлдберриз',
        'К перечислению Продавцу за реализованный Товар',
        'Количество доставок', 'Количество возврата',
        'Услуги по доставке товара покупателю',
        'Дата начала действия фиксации', 'Дата конца действия фиксации',
        'Признак услуги платной доставки', 'Общая сумма штрафов', 'Доплаты',
        'Виды логистики, штрафов и доплат', 'Стикер МП',
        'Наименование банка-эквайера', 'Номер офиса',
        'Наименование офиса доставки', 'ИНН партнера', 'Партнер', 'Склад',
        'Страна', 'Тип коробов', 'Номер таможенной декларации',
        'Номер сборочного задания', 'Код маркировки', 'ШК', 'Srid',
        'Возмещение издержек по перевозке/по складским операциям с товаром',
        'Организатор перевозки', 'Хранение', 'Удержания', 'Платная приемка',
        'chrtId', 'Фиксированный коэффициент склада по поставке'
    ]
    
    # Добавляем дополнительные столбцы для Unnamed
    for i in range(64, 71):
        headers.append(f'Unnamed: {i}')
    
    # Записываем заголовки
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Загружаем данные из базы
    financial_reports = FinancialReport.objects.all()
    for idx, report in enumerate(financial_reports, 2):
        ws.cell(row=idx, column=1, value=report.number)
        ws.cell(row=idx, column=2, value=report.delivery_number)
        ws.cell(row=idx, column=3, value=report.subject)
        ws.cell(row=idx, column=4, value=report.nomenclature_code)
        ws.cell(row=idx, column=5, value=report.brand)
        ws.cell(row=idx, column=6, value=report.supplier_article)
        ws.cell(row=idx, column=7, value=report.name)
        ws.cell(row=idx, column=8, value=report.size)
        ws.cell(row=idx, column=9, value=report.barcode)
        ws.cell(row=idx, column=10, value=report.document_type)
        ws.cell(row=idx, column=11, value=report.payment_basis)
        # Обработка дат с часовыми поясами
        order_date = report.order_date
        if order_date and hasattr(order_date, 'replace'):
            order_date = order_date.replace(tzinfo=None)
        ws.cell(row=idx, column=12, value=order_date)
        
        sale_date = report.sale_date
        if sale_date and hasattr(sale_date, 'replace'):
            sale_date = sale_date.replace(tzinfo=None)
        ws.cell(row=idx, column=13, value=sale_date)
        ws.cell(row=idx, column=14, value=report.quantity)
        ws.cell(row=idx, column=15, value=report.retail_price)
        ws.cell(row=idx, column=16, value=report.wb_sold_product)
        ws.cell(row=idx, column=17, value=report.agreed_product_discount)
        ws.cell(row=idx, column=18, value=report.promo_code_percent)
        ws.cell(row=idx, column=19, value=report.total_agreed_discount)
        ws.cell(row=idx, column=20, value=report.retail_price_with_discount)
        ws.cell(row=idx, column=21, value=report.kvv_reduction_rating)
        ws.cell(row=idx, column=22, value=report.kvv_reduction_promotion)
        ws.cell(row=idx, column=23, value=report.spp_discount)
        ws.cell(row=idx, column=24, value=report.kvv_size)
        ws.cell(row=idx, column=25, value=report.kvv_base_without_vat)
        ws.cell(row=idx, column=26, value=report.kvv_final_without_vat)
        ws.cell(row=idx, column=27, value=report.reward_before_services)
        ws.cell(row=idx, column=28, value=report.compensation_pickup_return)
        ws.cell(row=idx, column=29, value=report.acquiring_commission)
        ws.cell(row=idx, column=30, value=report.acquiring_commission_percent)
        ws.cell(row=idx, column=31, value=report.acquiring_payment_type)
        ws.cell(row=idx, column=32, value=report.wb_reward)
        ws.cell(row=idx, column=33, value=report.vat_wb_reward)
        ws.cell(row=idx, column=34, value=report.payment_to_seller)
        ws.cell(row=idx, column=35, value=report.delivery_count)
        ws.cell(row=idx, column=36, value=report.return_count)
        ws.cell(row=idx, column=37, value=report.delivery_services)
        # Обработка дат фиксации
        fixation_start = report.fixation_start_date
        if fixation_start and hasattr(fixation_start, 'replace'):
            fixation_start = fixation_start.replace(tzinfo=None)
        ws.cell(row=idx, column=38, value=fixation_start)
        
        fixation_end = report.fixation_end_date
        if fixation_end and hasattr(fixation_end, 'replace'):
            fixation_end = fixation_end.replace(tzinfo=None)
        ws.cell(row=idx, column=39, value=fixation_end)
        ws.cell(row=idx, column=40, value=report.paid_delivery_flag)
        ws.cell(row=idx, column=41, value=report.total_fines)
        ws.cell(row=idx, column=42, value=report.additional_payments)
        ws.cell(row=idx, column=43, value=report.logistics_types)
        ws.cell(row=idx, column=44, value=report.mp_sticker)
        ws.cell(row=idx, column=45, value=report.acquirer_bank)
        ws.cell(row=idx, column=46, value=report.office_number)
        ws.cell(row=idx, column=47, value=report.delivery_office)
        ws.cell(row=idx, column=48, value=report.partner_inn)
        ws.cell(row=idx, column=49, value=report.partner)
        ws.cell(row=idx, column=50, value=report.warehouse)
        ws.cell(row=idx, column=51, value=report.country)
        ws.cell(row=idx, column=52, value=report.box_type)
        ws.cell(row=idx, column=53, value=report.customs_declaration)
        ws.cell(row=idx, column=54, value=report.assembly_task)
        ws.cell(row=idx, column=55, value=report.marking_code)
        ws.cell(row=idx, column=56, value=report.sku)
        ws.cell(row=idx, column=57, value=report.srid)
        ws.cell(row=idx, column=58, value=report.transport_compensation)
        ws.cell(row=idx, column=59, value=report.transport_organizer)
        ws.cell(row=idx, column=60, value=report.storage)
        ws.cell(row=idx, column=61, value=report.deductions)
        ws.cell(row=idx, column=62, value=report.paid_reception)
        ws.cell(row=idx, column=63, value=report.chrt_id)
        ws.cell(row=idx, column=64, value=report.warehouse_coefficient)
        
        # Дополнительные столбцы
        for i in range(65, 71):
            ws.cell(row=idx, column=i, value=None)
    
    print(f"  ✅ Загружено {len(financial_reports)} записей финансовых отчетов")

def setup_sales_analysis_sheet(wb):
    """Настраивает лист 'Анализ продаж'"""
    print("📊 Настройка листа 'Анализ продаж'...")
    
    ws = wb['Анализ продаж']
    
    # Заголовки столбцов (48 столбцов)
    headers = [
        'Бренд', 'Предмет', 'Артикул', 'Размер', 'Баркод',
        'В пути до клиента', 'В пути от клиента', 'На складах',
        'ИТОГО остаток на ВБ', 'Оборачиваемость, дней (для отчетов за НЕДЕЛЮ!)',
        'Заказы, шт', 'Отказы', 'Продажи, шт', 'Возвраты, шт',
        'Продажи минус возвраты', 'Процент выкупа', 'Возврат, шт',
        'Логистика за возвраты', 'Продажи по ценам до СПП',
        'Продажи по ценам с СПП', 'Продажи за вычетом комиссии',
        'Продажи за вычетом комиссии (без учета возвратов)',
        'Возвраты', 'Продажи минус возвраты.1', 'Комиссия, руб',
        'Комиссия %', 'Логистика', 'Логистика на 1 продажу',
        'Эквайринг', 'Штраф', 'Доплаты', 'Компенсация подмен',
        'Возмещение брака', 'Средний чек', 'Себестоимость 1 шт',
        'Себестоимость проданного товара', 'Маржа до налогов',
        'Налог 6%', 'Маржа после налогов, руб', 'Маржа на 1 продажу, руб',
        'Маржинальность, %', 'ROI от себестоимости, %', 'GMROI, %',
        'Доля от общей выручки, %', 'Доля от общей маржи, %',
        'По себестоимости', 'По цене за вычетом комиссии',
        'По средней марже'
    ]
    
    # Записываем заголовки
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Загружаем данные из базы
    sales_analysis = SalesAnalysis.objects.all()
    for idx, analysis in enumerate(sales_analysis, 2):
        ws.cell(row=idx, column=1, value=analysis.brand)
        ws.cell(row=idx, column=2, value=analysis.subject)
        ws.cell(row=idx, column=3, value=analysis.article)
        ws.cell(row=idx, column=4, value=analysis.size)
        ws.cell(row=idx, column=5, value=analysis.barcode)
        ws.cell(row=idx, column=6, value=analysis.in_transit_to_client)
        ws.cell(row=idx, column=7, value=analysis.in_transit_from_client)
        ws.cell(row=idx, column=8, value=analysis.in_warehouses)
        ws.cell(row=idx, column=9, value=analysis.total_stock_wb)
        ws.cell(row=idx, column=10, value=analysis.turnover_days)
        ws.cell(row=idx, column=11, value=analysis.orders)
        ws.cell(row=idx, column=12, value=analysis.rejections)
        ws.cell(row=idx, column=13, value=analysis.sales)
        ws.cell(row=idx, column=14, value=analysis.returns)
        ws.cell(row=idx, column=15, value=analysis.sales_minus_returns)
        ws.cell(row=idx, column=16, value=analysis.purchase_percentage)
        ws.cell(row=idx, column=17, value=analysis.supplier_return)
        ws.cell(row=idx, column=18, value=analysis.logistics_for_returns)
        ws.cell(row=idx, column=19, value=analysis.sales_before_spp)
        ws.cell(row=idx, column=20, value=analysis.sales_with_spp)
        ws.cell(row=idx, column=21, value=analysis.sales_minus_commission)
        ws.cell(row=idx, column=22, value=analysis.sales_minus_commission_no_returns)
        ws.cell(row=idx, column=23, value=analysis.return_amount)
        ws.cell(row=idx, column=24, value=analysis.sales_minus_returns_amount)
        ws.cell(row=idx, column=25, value=analysis.commission)
        ws.cell(row=idx, column=26, value=analysis.commission_percent)
        ws.cell(row=idx, column=27, value=analysis.logistics)
        ws.cell(row=idx, column=28, value=analysis.logistics_per_unit)
        ws.cell(row=idx, column=29, value=analysis.acquiring)
        ws.cell(row=idx, column=30, value=analysis.fine)
        ws.cell(row=idx, column=31, value=analysis.additional_payments)
        ws.cell(row=idx, column=32, value=analysis.substitution_compensation)
        ws.cell(row=idx, column=33, value=analysis.defect_compensation)
        ws.cell(row=idx, column=34, value=analysis.average_check)
        ws.cell(row=idx, column=35, value=analysis.cost_per_unit)
        ws.cell(row=idx, column=36, value=analysis.sold_goods_cost)
        ws.cell(row=idx, column=37, value=analysis.margin_before_tax)
        ws.cell(row=idx, column=38, value=analysis.tax_6_percent)
        ws.cell(row=idx, column=39, value=analysis.margin_after_tax)
        ws.cell(row=idx, column=40, value=analysis.margin_per_unit)
        ws.cell(row=idx, column=41, value=analysis.margin_percent)
        ws.cell(row=idx, column=42, value=analysis.roi_from_cost)
        ws.cell(row=idx, column=43, value=analysis.gmroi)
        ws.cell(row=idx, column=44, value=analysis.revenue_share)
        ws.cell(row=idx, column=45, value=analysis.margin_share)
        ws.cell(row=idx, column=46, value=analysis.abc_by_cost)
        ws.cell(row=idx, column=47, value=analysis.abc_by_price)
        ws.cell(row=idx, column=48, value=analysis.abc_by_margin)
    
    print(f"  ✅ Загружено {len(sales_analysis)} записей анализа продаж")

def setup_summary_sheet(wb):
    """Настраивает лист 'Сводный'"""
    print("📈 Настройка листа 'Сводный'...")
    
    ws = wb['Сводный']
    
    # Заголовки
    ws['B1'] = 'Средние показатели за периода'
    ws['E1'] = 'Расходы маркетплейса'
    
    # Данные и формулы
    ws['B2'] = 'Продаж за вычетом возвратов, шт'
    ws['C2'] = '=\'Анализ продаж\'!O3'
    ws['B3'] = 'Процент выкупа'
    ws['C3'] = '=\'Анализ продаж\'!P3'
    ws['B4'] = 'Средний чек для покупателя после СПП'
    ws['C4'] = '=\'Анализ продаж\'!Y3'
    
    ws['E2'] = 'Комиссия'
    ws['F2'] = '=\'Анализ продаж\'!Y3'
    ws['E3'] = 'Логистика'
    ws['F3'] = '=\'Анализ продаж\'!AA3+\'Анализ продаж\'!R3'
    ws['E4'] = 'Хранение'
    ws['F4'] = '=\'Финотчеты (вставить)\'!BH1'
    
    print("  ✅ Настроены формулы для сводного листа")

def setup_other_sheets(wb):
    """Настраивает остальные листы"""
    print("📋 Настройка остальных листов...")
    
    # Лист "Остатки (вставить)"
    ws = wb['Остатки (вставить)']
    headers = [
        'Бренд', 'Предмет', 'Артикул продавца', 'Размер вещи',
        'В пути до клиента', 'В пути от клиента', 'Итого по складам'
    ]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Лист "План по выкупам"
    ws = wb['План по выкупам']
    headers = ['Unnamed: 0', 'Unnamed: 1', 'Unnamed: 2', 'выкупы', 'изначальная поз', 'всего заказзаов']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Данные для плана по выкупам
    data = [
        [None, None, 1, 1, 9, 10],
        [None, None, 2, 2, 9, 11],
        [None, None, 3, 4, 9, 13]
    ]
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Лист "Лист10" - шаблон
    ws = wb['Лист10']
    # Копируем заголовки из финансовых отчетов
    financial_headers = [
        '№', 'Номер поставки', 'Предмет', 'Код номенклатуры', 'Бренд',
        'Артикул поставщика', 'Название', 'Размер', 'Баркод', 'Тип документа',
        'Обоснование для оплаты', 'Дата заказа покупателем', 'Дата продажи',
        'Кол-во', 'Цена розничная', 'Вайлдберриз реализовал Товар (Пр)',
        'Согласованный продуктовый дисконт, %', 'Промокод %',
        'Итоговая согласованная скидка', 'Цена розничная с учетом согласованной скидки',
        'Размер снижения кВВ из-за рейтинга, %', 'Размер снижения кВВ из-за акции, %',
        'Скидка постоянного Покупателя (СПП)', 'Размер кВВ, %',
        'Размер кВВ без НДС, % Базовый', 'Итоговый кВВ без НДС, %',
        'Вознаграждение с продаж до вычета услуг поверенного, без НДС',
        'Возмещение за выдачу и возврат товаров на ПВЗ',
        'Эквайринг/Комиссии за организацию платежей',
        'Размер комиссии за эквайринг/Комиссии за организацию платежей, %',
        'Тип платежа за Эквайринг/Комиссии за организацию платежей',
        'Вознаграждение Вайлдберриз (ВВ), без НДС',
        'НДС с Вознаграждения Вайлдберриз',
        'К перечислению Продавцу за реализованный Товар',
        'Количество доставок', 'Количество возврата',
        'Услуги по доставке товара покупателю',
        'Дата начала действия фиксации', 'Дата конца действия фиксации',
        'Признак услуги платной доставки', 'Общая сумма штрафов', 'Доплаты',
        'Виды логистики, штрафов и доплат', 'Стикер МП',
        'Наименование банка-эквайера', 'Номер офиса',
        'Наименование офиса доставки', 'ИНН партнера', 'Партнер', 'Склад',
        'Страна', 'Тип коробов', 'Номер таможенной декларации',
        'Номер сборочного задания', 'Код маркировки', 'ШК', 'Srid',
        'Возмещение издержек по перевозке/по складским операциям с товаром',
        'Организатор перевозки', 'Хранение', 'Удержания', 'Платная приемка',
        'chrtId', 'Фиксированный коэффициент склада по поставке'
    ]
    
    for col, header in enumerate(financial_headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Инструкции
    ws['A3'] = '1. Перейдите в раздел ВБ: Отчеты - Аналитика - Отчет по остаткам на складе'
    ws['A4'] = '2. ❗️Нажмите на кнопку "настройка таблицы" и активируйте галочки "артикул продавца" и "размер вещи". Скачайте отчет'
    ws['A5'] = '3. Выделите все данные из полученного отчета, скопируйте и вставьте в данный лист'
    
    # Лист "Лист11" - пустой
    ws = wb['Лист11']
    for row in range(1, 14):
        for col in range(1, 10):
            ws.cell(row=row, column=col, value=None)
    
    print("  ✅ Настроены все остальные листы")

def apply_formatting(wb):
    """Применяет форматирование ко всем листам"""
    print("🎨 Применение форматирования...")
    
    # Стили
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    # Применяем стили к заголовкам
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
    
    print("  ✅ Форматирование применено")

def main():
    """Основная функция"""
    print("🚀 Начало воссоздания Excel файла...")
    print("=" * 50)
    
    try:
        # Создаем структуру
        wb = create_excel_structure()
        
        # Настраиваем листы
        setup_nomenclature_sheet(wb)
        setup_financial_reports_sheet(wb)
        setup_sales_analysis_sheet(wb)
        setup_summary_sheet(wb)
        setup_other_sheets(wb)
        
        # Применяем форматирование
        apply_formatting(wb)
        
        # Сохраняем файл
        filename = f'Воссозданный_файл_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        wb.save(filename)
        
        print("=" * 50)
        print(f"✅ Excel файл успешно воссоздан: {filename}")
        print(f"📊 Листов: {len(wb.sheetnames)}")
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"  📋 {sheet_name}: {ws.max_row} строк x {ws.max_column} столбцов")
        
        print("🎉 Процесс завершен успешно!")
        
    except Exception as e:
        print(f"❌ Ошибка при создании файла: {e}")
        return False
    
    return True

if __name__ == "__main__":
    success = main()
    sys.exit(0 if success else 1)
